from subprocess import call
import win32com.client
from time import time
import os
import pandas as pd
import win32api
import time
from openpyxl import load_workbook
import openpyxl
import re
import datetime


# Get username PC
us = os.getlogin()

result = win32api.MessageBox(None,"Você está com o SAP aberto neste momento?", "Aviso",1)

if result == 1:
 print ('SAP Success Conected!')
elif result == 2:
 raise SystemExit


dataframe1 = pd.read_excel(r'C:\Users\%s\Desktop\Transeferncia_BH\Itens.xlsx' %us,dtype=str) 

print(dataframe1)

# SAP Conection
SapGuiAuto = win32com.client.GetObject("SAPGUI")
application = SapGuiAuto.GetScriptingEngine
connection = application.Children(0)
session = connection.Children(0)

dataframe1.describe()


# To get all SAP children objects and ids
def get_all_children(parentNodeObj:win32com.client.CDispatch):
    """
    Function to retrieves all child objects and their IDs for a given parent object recursively.
    """
    childrenList = []
    def sub_get_all_children(parentNodeObj):
        for obj in parentNodeObj.children:
            childrenList.append({'Obj': obj, 'Id': obj.id, 'Text': obj.text})
            try: objLen = len(obj.children)
            except: objLen = 0
            if objLen > 0:
                sub_get_all_children(obj)
    
    sub_get_all_children(parentNodeObj)

    return childrenList

def get_id(fieldName):
    idsList = [i['Id'] for i in allChildren if fieldName in i['Id']]
    if len(idsList) == 0:
      print('Nenhum objeto encontrado!')
      os.system('pause')
    else:
        return [i['Id'] for i in allChildren if fieldName in i['Id']][0]
    
# Change language to EN 
session.findById("wnd[0]/tbar[0]/okcd").text = "zca_chng_lgn_languag"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/usr/ctxtP_SPRAS").text = "EN"
session.findById("wnd[0]/tbar[1]/btn[8]").press()

# Enter in ME21N
session.findById("wnd[0]/tbar[0]/okcd").text = "/nme21n"
session.findById("wnd[0]").sendVKey (0)
# header in ME21N
allChildren = get_all_children(session.findById("wnd[0]/usr/"))
session.findById(get_id('TOPLINE-BSART')).key = "UB"
allChildren = get_all_children(session.findById("wnd[0]/usr/"))
session.findById(get_id('TOPLINE-SUPERFIELD')).text = "4403"
session.findById(get_id('ctxtMEPO1222-EKORG')).text = "4425"
session.findById(get_id('ctxtMEPO1222-EKGRP')).text = "BR9"
session.findById(get_id('ctxtMEPO1222-BUKRS')).text = "4400"
session.findById(get_id('btnDYN_4000-BUTTON')).press()


# Loop with itens from worksheet
allChildren = get_all_children(session.findById("wnd[0]/usr/"))
for index, row in dataframe1.iterrows():

    # Insert itens
    session.findById(get_id('ctxtMEPO1211-EMATN[4,1]')).text = row['Item']
    # Insert qnt
    session.findById(get_id('txtMEPO1211-MENGE[6,1]')).text = row['Qnt']
    # Insert plant destiny
    session.findById(get_id('ctxtMEPO1211-NAME1[11,1]')).text = "4425"
    # Insert storage location
    session.findById(get_id('ctxtMEPO1211-LGOBE[12,1]')).text = "0001"
    # Insert batch
    session.findById(get_id('ctxtMEPO1211-CHARG[13,1]')).text = row['Lote']

    # Get scrollbar position
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    scrollbar_position = session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position
    
    # Verify if position scrollbar change
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    new_scrollbar_position = session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position
    if new_scrollbar_position == scrollbar_position:
        # Roll for down the Scrollbar
        session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position += 1

    # Insert qnt again
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))    
    session.findById(get_id('txtMEPO1211-MENGE[6,0]')).text = row['Qnt']
 
    # Jump line
    session.findById("wnd[0]").sendVKey(0)
    
    # Wait Process SAP
    time.sleep(1)

    
    # Verify if position scrollbar change
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    new_scrollbar_position = session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position
    if new_scrollbar_position == scrollbar_position:
        # Roll for down the Scrollbar
        session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position += 1


# Finish loop rolling to up scrollbar
while session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position > 0:
    session.findById(get_id('tblSAPLMEGUITC_1211')).verticalScrollbar.position -= 1

# Count itens worksheet
total_itens = len(dataframe1['Item'])
print(f'Total de itens da planilha - {total_itens}')

# loop to delete vendor batch and delete route
for i in range(1, total_itens + 1):
    # Create a variable from count itens 
    novo_valor = " " * (4 - len(str(i))) + str(i)
    session.findById(get_id('cmbDYN_6000-LIST')).key = novo_valor
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    session.findById(get_id('tabpTABIDT3')).select()
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    session.findById(get_id('txtMEPO1319-LICHA')).text = ""

    # Delete Route
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))    
    session.findById(get_id('tabpTABIDT19')).select()
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    session.findById(get_id('ctxtMEPO1331-TRAGR')).text = ""
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    session.findById(get_id('ctxtMEPO1331-ROUTE')).text = ""
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]").sendVKey (0)

    

# Save purchase order
session.findById("wnd[0]/tbar[0]/btn[11]").press()
session.findById("wnd[1]/usr/btnSPOP-VAROPTION1").press()

# Read the Excel file
file_path = r'C:\Users\%s\Desktop\Transeferncia_BH\Itens.xlsx' % us
workbook = openpyxl.load_workbook(file_path)

# Access the 'Registros' tab in workbook
sheet = workbook['Registros']

# Text from the SAP session
text_from_sap = session.findById("wnd[0]/sbar").Text

# Use a regular expression to extract numbers
numbers = re.findall(r'\d+', text_from_sap)

# If numbers are found, concatenate them into a single string
if numbers:
    extracted_numbers = ''.join(numbers)
else:
    extracted_numbers = ''

# Find the next blank row in column A of the worksheet
next_row = len(sheet['A']) + 1

# Put the extracted numbers from SAP in the next blank row in column "Pedido"
sheet.cell(row=next_row, column=1, value=extracted_numbers)

# Save the workbook
workbook.save(file_path)

# Select tab "Registro" in workbook
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Registros']
column_pedido = []
for cell in sheet['A'][1:]:
    column_pedido.append(cell.value)
dataframe2 = pd.DataFrame({'Pedido': column_pedido})

# Enter in VL10B
for pedido in dataframe2['Pedido']:
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nVL10B"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/tabsTABSTRIP_ORDER_CRITERIA/tabpS0S_TAB5").select()
    session.findById("wnd[0]/usr/ctxtST_LEDAT-LOW").text = ""
    session.findById("wnd[0]/usr/ctxtST_LEDAT-HIGH").text = ""
    session.findById("wnd[0]/usr/ctxtP_LERUL").text = ""
    allChildren = get_all_children(session.findById("wnd[0]/usr/"))
    session.findById(get_id('ctxtST_EBELN-LOW')).text = pedido
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("AMPEL")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("AUART")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("VBELN")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("VBELV")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("KUNWE")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("WADAT")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("KUNNR")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("MANDT")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("LPRIO")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("VSTEL")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("LIFSP")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("VKGRP")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn ("VKBUR")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectedRows = ("0")
    session.findById("wnd[0]/tbar[1]/btn[19]").press()
    session.findById("wnd[0]/tbar[1]/btn[45]").press()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").select()   
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

# Trata clipboard com número da delivery - sendo SD Doc o numero da delivery
gusta = pd.read_clipboard(sep='|', header=3, dtype=str)
gusta = gusta[gusta['   SD Doc.'].notnull()][['   SD Doc.']]
gusta = gusta[gusta['   SD Doc.']!='          ']

# Read the Excel file
file_path = r'C:\Users\%s\Desktop\Transeferncia_BH\Itens.xlsx' % us
workbook = openpyxl.load_workbook(file_path)

# Access the 'Registros' tab in workbook
sheet = workbook['Registros']

# Find the next blank row in column B in worksheet
next_row = 1
while sheet.cell(row=next_row, column=2).value is not None:
    next_row += 1

sd_doc_value = gusta.iloc[0, 0]
# Put the extracted numbers from SAP in the next blank row in column "Pedido"
sheet.cell(row=next_row, column=2, value=sd_doc_value)

# Save the workbook
workbook.save(file_path)

# Query para obter as quantidades de itens geradas do pedido
session.findById("wnd[0]/tbar[0]/okcd").text = "/nsq00"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/tbar[1]/btn[19]").press()
session.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").currentCellRow = -1
session.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").selectColumn ("DBGBNUM")
session.findById("wnd[1]/tbar[0]/btn[29]").press()
session.findById("wnd[2]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = "SD_LO"
session.findById("wnd[2]/tbar[0]/btn[0]").press()
session.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").selectedRows = ("0")
session.findById("wnd[1]/usr/cntlGRID1/shellcont/shell").doubleClickCurrentCell()
session.findById("wnd[0]/usr/ctxtRS38R-QNUM").text = "ZSD_LO_11"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/ctxtSP$00001-LOW").text = sd_doc_value
session.findById("wnd[0]/usr/ctxt%LAYOUT").text = "/gustabh"
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem ("&PC")
session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").select()
session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").setFocus()
session.findById("wnd[1]/tbar[0]/btn[0]").press()

# trata o output do clipboard
gusta2 = pd.read_clipboard(sep='|', header=6, dtype=str)

# Remova espaços extras antes e depois dos nomes das colunas
gusta2.columns = gusta2.columns.str.strip()

# Selecione a coluna "Delivery quantity" sem linhas em branco
gusta2 = gusta2[gusta2['Delivery quantity'].notnull()][['Batch', 'Delivery quantity']]

# Inicialize o contador de materiais
material_counts = 0

# Loop para iterar sobre os itens e fazer o input da quantidade no Picking
for index, row in gusta.iterrows():
    
    # VL02N saída de mercadorias na delivery
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nvl02n"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtLIKP-VBELN").text = sd_doc_value
    session.findById("wnd[0]").sendVKey(0)
    try:
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02").select() # Select Picking table
    except:
        print("Aba Picking já selecionada!")
    

# Inicialize o contador de materiais
material_counts = 0

# Inicie o loop externo
index = 0 
encerrar_loop = False  # Variável de controle para o loop externo
while index < len(gusta2) and not encerrar_loop:
    row = gusta2.iloc[index]

    # Determine o número da linha com base na contagem da coluna "Batch"
    linha_batch = material_counts

    # Loop para iterar sobre os itens e fazer o input da quantidade no Picking
    diabo = 0
    for item_index in range(total_itens):

        try:
            Var = session.findById(f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02/ssubSUBSCREEN_BODY:SAPMV50A:1104/tblSAPMV50ATC_LIPS_PICK/ctxtLIPS-MATNR[1,{diabo}]").text
        except:
            session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02/ssubSUBSCREEN_BODY:SAPMV50A:1104/tblSAPMV50ATC_LIPS_PICK").verticalScrollbar.position =  item_index
            diabo = 0

        
        codigo = f"[6,{diabo}]"
        codigo_texto = f'wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02/ssubSUBSCREEN_BODY:SAPMV50A:1104/tblSAPMV50ATC_LIPS_PICK/txtLIPSD-PIKMG{codigo}'
        campo_id = f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02/ssubSUBSCREEN_BODY:SAPMV50A:1104/tblSAPMV50ATC_LIPS_PICK/txtLIPSD-G_LFIMG[4,{diabo}]"
        campo = session.findById(campo_id)
        texto_do_campo = campo.Text
        session.findById(codigo_texto).Text = texto_do_campo

        diabo += 1

        
    encerrar_loop = True
    break

# Access the 'Histórico' tab in workbook
historico_sheet = workbook['Historico']

# Find the next blank row in column A in 'Histórico' worksheet
next_historico_row = 1
while historico_sheet.cell(row=next_historico_row, column=1).value is not None:
    next_historico_row += 1

# Move data from 'Registros' to 'Histórico' (excluding the header row)
header_row = 1  
for row in sheet.iter_rows(min_row=header_row+1, max_row=next_row, min_col=1, max_col=3):
    historico_row = next_historico_row
    for cell in row:
        historico_sheet.cell(row=historico_row, column=cell.column, value=cell.value)
    next_historico_row += 1

# Clear the data from 'Registros' (excluding the header row)
for row in sheet.iter_rows(min_row=header_row+1, max_row=next_row, min_col=1, max_col=3):
    for cell in row:
        cell.value = None

# Save the workbook again to include the updated 'Histórico' sheet
workbook.save(file_path)

# SM register in VL02N
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07").select()
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07/ssubSUBSCREEN_BODY:SAPMV50A:2110/cmbLIKP-LIFSK").key = " "
session.findById("wnd[0]/tbar[0]/btn[3]").press()
session.findById("wnd[0]/tbar[1]/btn[20]").press()

data_atual = datetime.datetime.now()
data_formatada = data_atual.strftime("%d.%m.%Y")

# Wait delivery load
time.sleep(15)

# J1BNFE NF flag
session.findById("wnd[0]/tbar[0]/okcd").text = "/nj1bnfe"
session.findById("wnd[0]").sendVKey (0)
session.findById("wnd[0]/usr/ctxtDATE0-LOW").text = data_formatada
session.findById("wnd[0]/usr/ctxtBUKRS-LOW").text = "4400"
session.findById("wnd[0]/usr/txtSHIPT-LOW").text = "4403"
session.findById("wnd[0]/usr/txtUSERCRE-LOW").text = us
session.findById("wnd[0]/tbar[1]/btn[8]").press()
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_VARIANT")
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").selectContextMenuItem ("&LOAD")
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").currentCellRow = (5)
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").selectedRows = "5"
session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").setCurrentCell (-1,"")
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").selectColumn ("STATUS")
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").selectColumn ("ACTION_REQU")
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").selectColumn ("ERRLOG")
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").selectColumn ("DOCNUM")
session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").selectedRows = ("0")
session.findById("wnd[0]/tbar[1]/btn[30]").press()